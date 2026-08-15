"""Content and writer for the 'Sources & notes' tab."""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

H1 = Font(bold=True, size=14, color="1F3864")
H2 = Font(bold=True, size=11, color="FFFFFF")
KEY = Font(bold=True, size=10)
BODY = Font(size=10)
MONO = Font(size=9, color="1F4E79")
FILL2 = PatternFill("solid", fgColor="1F3864")
BAND = PatternFill("solid", fgColor="F2F5FA")

SECTIONS = [
    ("ELECTRICITY TAB", [
        ("What it is", "Daily wholesale (spot) electricity prices for every National Electricity Market region, from the start of the NEM in December 1998. Units are $/MWh, GST exclusive."),
        ("Publisher", "Australian Energy Market Operator (AEMO) - 'Aggregated price and demand data'."),
        ("Landing page", "https://www.aemo.com.au/energy-systems/electricity/national-electricity-market-nem/data-nem/aggregated-data"),
        ("Files used", "https://www.aemo.com.au/aemo/data/nem/priceanddemand/PRICE_AND_DEMAND_<YYYYMM>_<REGION>.csv - one file per region per month, from December 1998 to the current month."),
        ("Source fields", "REGION, SETTLEMENTDATE, TOTALDEMAND (MW), RRP (regional reference price, $/MWh), PERIODTYPE."),
        ("Interval length", "30-minute trading intervals to 30 September 2021, then 5-minute intervals from 1 October 2021 under five-minute settlement. The 'intervals' column records how many intervals went into each day (48 before the change, 288 after) so partial days are visible."),
        ("Regions over time", "Not every region existed for the whole period. Tasmania (TAS1) joined the NEM in May 2005, so its columns are blank before then. The Snowy region (SNOWY1) was a NEM region until it was abolished after June 2008, so its columns are blank afterwards. Both are real history and are left in rather than trimmed. A 404 from AEMO for a region-month is recorded as did-not-exist, not as a download failure."),
        ("Date convention", "AEMO stamps each interval with its interval-ENDING time. One second is subtracted before taking the calendar date, so a day runs 00:00-24:00 local market time (AEST, no daylight saving) and the interval ending at midnight is credited to the day it belongs to."),
        ("Derived columns", "price_dwa = demand-weighted average price, sum(RRP x TOTALDEMAND) / sum(TOTALDEMAND). price_mean = simple arithmetic mean across intervals. price_min / price_max = lowest and highest interval price in the day. demand_mean_mw = average regional demand."),
        ("Which price to use", "price_dwa is the better headline number - it is what load actually paid on average. price_mean over-weights low-demand overnight intervals. The two diverge a lot on volatile days."),
        ("Known quirks", "Negative prices and prices at the market price cap are genuine, not data errors; they are frequent in SA and increasingly in VIC/QLD. Extreme daily maxima usually reflect a small number of capped intervals. Note the market price cap itself has been raised many times since 1998, so early and late extremes are not on the same scale."),
        ("Coverage limits", "NEM only. Western Australia (WEM/SWIS) and the Northern Territory are separate markets and are NOT in this tab."),
        ("Update behaviour", "AEMO adds the current month's file continuously; the current month is partial until it ends. Prices are as-published and AEMO may revise them."),
    ]),
    ("GAS TAB", [
        ("What it is", "Daily east coast wholesale gas market prices, in $/GJ, for the four AEMO-operated gas markets: Victoria (DWGM) and the Sydney, Adelaide and Brisbane Short Term Trading Market hubs."),
        ("Publisher", "AEMO."),
        ("Victoria - file", "https://www.aemo.com.au/-/media/files/gas/dwgm/dwgm-prices-and-demand.xlsx  (linked from the 'VIC wholesale price & withdrawals' page). Covers 1 February 2007 onward."),
        ("Victoria - method", "The Declared Wholesale Gas Market re-schedules five times per gas day, at 6am, 10am, 2pm, 6pm and 10pm. The source file gives a price for each schedule. VIC_DWGM_6am is the 6am (start of gas day) schedule price - the usual headline number. VIC_DWGM_schedule_mean is the unweighted mean of that day's five schedule prices. VIC_DWGM_n_schedules shows how many schedule prices AEMO actually published for that day, so the mean is never silently computed off a short day - it is 5 on every day in this extract except 16 November 2015, where the source carries only 4."),
        ("STTM - file", "https://www.aemo.com.au/-/media/files/gas/sttm/data/sttm-price-and-withdrawals.xlsx  (linked from the 'Daily STTM reports' page)."),
        ("STTM - method", "One ex-ante and one ex-post price per hub per gas day, taken directly from the file - no aggregation. Ex-ante is the price set the day before from scheduled quantities and is the price normally quoted. Ex-post is recalculated after the fact from actual flows. Source history starts 1 Sep 2010 for Sydney and Adelaide and 1 Dec 2011 for Brisbane, all well before this extract's 2015 start."),
        ("Coverage", "Starts 1 February 2007, the first gas day in the AEMO DWGM master file, and runs to the end of the last completed month. The Victorian columns cover that whole span with a value on every calendar day. The STTM hub columns start later because the market did: Sydney and Adelaide from 1 September 2010, Brisbane from 1 December 2011. Blanks before those dates mean the hub did not exist yet."),
        ("Publication lag", "Both AEMO workbooks are refreshed monthly, so the gas tab typically ends at the close of the last completed month and runs a few weeks behind the electricity and petrol tabs. This is a property of the source, not a gap in the extract."),
        ("Not included", "The Wallumbilla / South East Queensland Gas Supply Hub. AEMO publishes GSH trade files only on a short rolling window on NEMWeb; no back-to-2015 master file was found, so rather than guess, it is left out. It could be built up going forward from the rolling files if you want it."),
        ("Coverage limits", "Wholesale market prices only - not retail gas bills, not LNG netback, not contract prices. Western Australia and the Northern Territory have no equivalent spot market and are not covered."),
    ]),
    ("PETROL TGP TAB (wholesale)", [
        ("What it is", "Daily wholesale terminal gate prices in cents per litre, GST inclusive, for petrol and diesel across seven capitals plus a national average. This is the cost of fuel at the refinery/import terminal gate - it is not what motorists pay."),
        ("Publisher", "Australian Institute of Petroleum (AIP), data prepared for AIP by ORIMA Research."),
        ("Page", "https://www.aip.com.au/historical-ulp-and-diesel-tgp-data  (the workbook filename carries its publication date and changes weekly, so the update script scrapes the current link rather than hard-coding it)."),
        ("Method", "AIP's stated method: the average of the terminal gate prices published by BP Australia, Ampol, Viva Energy and ExxonMobil for that day, inclusive of GST, collated each weekday morning. The national average is weighted by volume sold in each state (source: Australian Petroleum Statistics). Taken as published - no recalculation here."),
        ("Gaps", "Collated on business days only, so weekend dates are absent entirely. Verified: 3,032 rows for 2015-01-01 to 2026-08-14, all Monday-Friday."),
        ("Old note", "AIP notes that Brisbane prices before 11 July 2009 exclude the former Queensland fuel subsidy. That is before this extract's 2015 start and does not affect it."),
        ("Why it is separate", "TGP is a clean, complete, national weekday series. Retail is patchier and state-by-state, so the two are kept on separate tabs rather than interleaved with mismatched blanks."),
        ("Relationship to retail", "On level the two track closely - Perth retail ULP correlates 0.945 with Perth TGP, at an average margin of 13.1 c/L. On daily MOVEMENT they are all but unrelated (correlation of daily changes 0.05), because retail is driven by the discount cycle rather than by wholesale cost. Use TGP for cost and policy questions; use retail only when the question is actually about pump prices or retail margin."),
    ]),
    ("PETROL RETAIL TAB", [
        ("What it is", "Daily average retail pump prices in cents per litre, GST inclusive, for the three states that publish usable price data: New South Wales, Queensland and Western Australia. WA is given twice - all-of-state and Perth metro."),
        ("Coverage differs by state", "WA from 2015-01-01 (the start of this extract; FuelWatch itself goes back to 2001 and could be extended on request). NSW from 1 August 2016. QLD reporting starts December 2018 but the first unleaded day that clears the 5-site minimum is 1 February 2019. Earlier dates are blank for a state simply because the scheme did not exist yet. The _sites columns show how many sites underpin each day's average - check them before trusting a thin early period."),
        ("WA - source", "FuelWatch, Government of Western Australia. Monthly station-level CSVs enumerated via the FuelWatch API (/api/report/monthly-retail-prices). Each row is one site, one product, one day."),
        ("WA - method", "Straight daily mean of every reported site price. WA's price notification scheme fixes posted prices for the day, which is why the WA series is genuinely daily and unusually clean."),
        ("NSW - source", "FuelCheck, NSW Fair Trading, via data.nsw.gov.au (dataset a97a46fc-2bdd-4b90-ac7f-0cb1e8d7ac3b). Monthly spreadsheets of price-change events."),
        ("QLD - source", "Queensland Fuel Price Reporting, via data.qld.gov.au ('Fuel price reporting' datasets, monthly 'changes only' CSVs). Prices are published in tenths of a cent and are divided by 10 here; timestamps are UTC and are converted to AEST before the date is taken."),
        ("NSW/QLD - method", "Both states publish price CHANGES, not daily snapshots, so a daily average requires carrying each site's last posted price forward until it changes. A site is dropped once it has not reported for 30 days, so closed sites stop dragging the average, and a day is only written when at least 5 sites are reporting that product."),
        ("NSW - known gaps", "August 2023, July 2024, March 2025 and July 2025 are simply not published on data.nsw.gov.au - verified as absent from the portal, not a parsing failure. Those months are blank, and the following month starts with a reduced site count while the carried-forward prices refresh."),
        ("Warm-up at the start", "For NSW and QLD the first week or two of each state's series rests on fewer sites than normal, because the forward-fill only knows about a site once it has posted its first price change. Site counts climb quickly to their steady level - NSW opens at ~474 sites and settles well above that. Treat the opening fortnight of each state as indicative and watch the _sites column."),
        ("Publication lag", "WA is published continuously and runs to within a day or two of the build date. NSW and Queensland publish one file per month, so their columns stop at the end of the last completed month and lag WA by up to five weeks. Verified at this build: WA to 2026-08-14, NSW and QLD both to 2026-07-31."),
        ("Key caveat", "These are UNWEIGHTED averages across sites, not volume-weighted. ACCC and AIP retail figures are volume-weighted and will not match these exactly. Use this tab for shape, spread and direction, not as an official retail benchmark."),
        ("Statewide, not city", "NSW and QLD figures are statewide, mixing metropolitan and regional sites (regional fuel is typically dearer). Only WA is split, because FuelWatch publishes an explicit Metro region flag; no postcode-to-city boundary has been invented for the other states."),
        ("Price cycles", "Retail prices in the eastern capitals move on discount cycles of several weeks, and Perth on a strict weekly cycle set by WA's 24-hour rule. Day-to-day movement in this tab is mostly cycle, not cost. Averaging to weekly or monthly removes most of it."),
        ("Not covered", "Victoria, South Australia, Tasmania, the ACT and the Northern Territory. Victoria has no reporting scheme; SA and Tasmania publish no open historical dataset that could be found; the NT's MyFuel data is incomplete and stops in November 2024. Rather than pad the tab with partial or stale series, they are left out."),
    ]),
    ("DAILY / MONTHLY / QUARTERLY / ANNUAL TABS", [
        ("Structure", "Every series appears four times - daily, monthly, quarterly and annual. The daily tab is the source of truth; the monthly, quarterly and annual tabs are averages of it. Columns are identical across the four frequencies so a formula written against one works against the others."),
        ("Averaging rule", "Monthly, quarterly and annual figures are simple means of the daily values present in that period. Days with no value are skipped, never treated as zero, so a source gap pulls the average toward the days that do exist rather than dragging it down."),
        ("Electricity is different", "Electricity monthly, quarterly and annual demand-weighted prices are accumulated from the underlying 5-minute and 30-minute intervals, NOT averaged from the daily figures. An average of daily averages is not a demand-weighted average and would misprice volatile months, where a few high-demand, high-price intervals do most of the work. The min and max columns are likewise the true extremes over the period, not averages of daily extremes."),
        ("n_days column", "Every monthly, quarterly and annual row carries n_days - how many days that average actually rests on. Use it to spot partial periods. The current month, current quarter and current year are always partial, and periods with a source gap will read low."),
        ("Count columns", "Interval and schedule counts are summed over the period. Site counts are averaged, since a sum of daily site counts is meaningless."),
        ("Calendar, not financial", "Quarters are calendar quarters, labelled YYYY-Qn: Q1 is January-March. Annual means calendar year. Financial years and financial-year quarters (where Q1 is July-September) are not included; they can be added if wanted."),
        ("Validation", "The annual petrol TGP figures were checked against AIP's own published calendar-year averages across 352 city-year-fuel cells. Maximum difference 0.07 c/L, with 279 of 352 within 0.01 c/L. The small residual is explained by AIP publishing its daily values rounded to one decimal place."),
    ]),
    ("SUMMARY CHARTS TAB", [
        ("What it is", "Twenty-eight native Excel charts - every series at daily, monthly, quarterly and annual frequency, plus four two-scale electricity-and-gas charts, one per state. They are real charts pointing at cells on the data tabs, not pictures, so they redraw whenever the workbook is rebuilt and can be copied straight into a document or deck."),
        ("Daily charts are windowed", "Each series has two daily charts - the last two years and the last five - and petrol has two more, one for 2026 and one for 2025. Those two carry tighter axes of their own - 125-350 c/L for 2026, 125-200 for 2025 - so a single year's movement is legible instead of being flattened into the full band. The two years are therefore NOT on a common scale: compare the shapes, and read the numbers off the axis before comparing levels between them. The current-year chart is labelled with the date it runs to, because it is a part year. Full daily history is on the daily tabs; twenty-eight years of daily spot prices in a chart this size is a solid block of ink. The windows are measured back from each series' own last date, so the gas ones end a few weeks earlier than the rest."),
        ("Petrol charts have a fixed scale", "The petrol TGP and petrol retail charts are pinned to 75-350 c/L rather than auto-scaled (except the calendar-year charts, which have their own tighter axes), so wholesale and retail, and every frequency, can be read against one another without rescaling by eye. The band was set wide enough to contain every charted petrol value, including the April 2020 TGP trough at 80.8 c/L and the April 2026 peak at 325.9, so nothing is drawn off the edge of a plot. The build checks this on every run and stops with an error rather than quietly drawing a series outside its axis, so if a future refresh moves past either end the band gets widened rather than the data hidden. No petrol chart crops anything: 2026 has to reach 350 because TGP diesel peaked at 325.9 c/L on 9 April 2026, and the build reports any chart whose axis would cut a value off. Electricity and gas are left auto-scaled; their spikes are order-of-magnitude, and a fixed range would either flatten them or cut them off."),
        ("Colour is consistent", "A series keeps the same colour across every chart it appears in, so NSW is the same blue on the daily, monthly, quarterly and annual electricity charts. The palette was checked for colour-vision-deficiency separation."),
        ("What is charted", "Electricity: demand-weighted price for the five current regions. SNOWY1 is left off - it ceased to be a region in July 2008 and would be a dead line across four fifths of every chart. Gas: VIC DWGM 6am plus the three STTM hubs ex-ante. Petrol: the national TGP figures, and retail ULP by state. The other columns are all on the data tabs."),
        ("The two-scale charts", "There is one per state: each pairs that state's electricity spot price ($/MWh, left axis) with its own gas hub ($/GJ, right axis) - NSW with Sydney STTM, Victoria with the DWGM, Queensland with Brisbane STTM, South Australia with Adelaide STTM - so both lines describe one market rather than two unrelated ones. Each starts where its gas hub starts, which is why the windows differ: 2007-Q1 for Victoria, 2010-Q3 for Sydney and Adelaide, 2011-Q4 for Brisbane. Tasmania has no chart here - it is in neither the STTM nor the DWGM, so there is no Tasmanian gas price to pair with; TAS spot prices are on the electricity tabs. Read the shapes and their timing against each other, NOT the crossings: where the two lines sit relative to one another is set by the two axis ranges, not by anything in the market. Two scales on one chart always flatter a relationship, so treat it as a prompt to go and check the numbers, not as evidence on its own."),
    ]),
    ("GENERAL", [
        ("Units summary", "Electricity $/MWh (GST exclusive). Gas $/GJ. Petrol cents per litre (GST inclusive), both wholesale and retail. All prices nominal - no CPI adjustment has been applied."),
        ("Blank cells", "A blank means the source published no value for that date - weekends for TGP, or dates after a source's last refresh. Blanks are never filled, interpolated or carried forward."),
        ("No estimated data", "Every number in this workbook comes from a published source file. Nothing is modelled, back-filled or estimated."),
        ("Rebuilding / updating", "Type PRICES in a terminal. It re-downloads what can have changed and rebuilds this workbook in place, in about five minutes. PRICES NO RETAIL does the same but skips the retail step and omits the retail tab. Scripts live in ~/Documents/Work/au_fuel_prices/."),
        ("Caching", "Closed months never change, so monthly aggregates are cached on disk and only the current and previous month are re-fetched. The first build downloads a few GB; reruns download a few MB. Deleting the work/ directory forces a full rebuild."),
    ]),
]


def write_notes(wb, generated, coverage, retail=True):
    ws = wb.create_sheet("Sources & notes", 0)
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 132
    r = 1
    ws.cell(row=r, column=1, value="Australian daily energy & fuel prices - sources and method").font = H1
    r += 1
    ws.cell(row=r, column=1, value=f"Generated {generated}. Every series below is built directly from the published source file named in that section.").font = Font(italic=True, size=9, color="595959")
    r += 2
    ws.cell(row=r, column=1, value="Coverage at build time").font = KEY
    r += 1
    for tab, desc in coverage:
        ws.cell(row=r, column=1, value=tab).font = BODY
        ws.cell(row=r, column=2, value=desc).font = BODY
        r += 1
    r += 1
    sections = [x for x in SECTIONS if retail or not x[0].startswith('PETROL RETAIL')]
    for title, items in sections:
        c = ws.cell(row=r, column=1, value=title)
        c.font, c.fill = H2, FILL2
        c2 = ws.cell(row=r, column=2, value="")
        c2.fill = FILL2
        ws.row_dimensions[r].height = 20
        c.alignment = Alignment(vertical="center")
        r += 1
        for i, (k, v) in enumerate(items):
            a = ws.cell(row=r, column=1, value=k)
            a.font, a.alignment = KEY, Alignment(vertical="top", wrap_text=True)
            b = ws.cell(row=r, column=2, value=v)
            b.font = MONO if v.startswith("http") else BODY
            b.alignment = Alignment(vertical="top", wrap_text=True)
            if i % 2:
                a.fill = b.fill = BAND
            ws.row_dimensions[r].height = max(15, 13 * (len(v) // 118 + 1))
            r += 1
        r += 1
    ws.sheet_view.showGridLines = False
    return ws
