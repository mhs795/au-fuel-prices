"""Assemble the daily, monthly, quarterly and annual series into one formatted Excel
workbook, plus a Summary charts tab of native Excel charts drawn from those tabs.

  build_workbook.py <work_dir> <out.xlsx> [--no-retail]

Reads {elec,gas,tgp,retail}_{daily,monthly,quarterly,annual}.csv from the work directory. The
retail tabs are omitted entirely when --no-retail is given.
"""
import csv, os, sys, datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from notes import write_notes
from build_charts import write_charts

WORK, OUT = sys.argv[1], sys.argv[2]
RETAIL = "--no-retail" not in sys.argv[3:]

HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
TITLE_FONT = Font(bold=True, size=13, color="1F3864")
SUB_FONT = Font(italic=True, size=9, color="595959")
THIN = Side(style="thin", color="BFBFBF")
COUNT_SUFFIXES = ("_intervals", "_n_schedules", "_sites", "n_days", "_qty_GJ")

GROUPS = [
    ("elec", "Electricity", "Electricity — NEM regional spot prices", "$/MWh, GST exclusive. "
     "dwa = demand-weighted average. Source: AEMO aggregated price and demand data.", "0.00"),
    ("gas", "Gas", "Gas — east coast wholesale market prices", "$/GJ. Victorian DWGM schedule "
     "prices, STTM hub ex-ante/ex-post prices, and Gas Supply Hub traded prices and "
     "volumes at Wallumbilla and SEQ (2026 on). Source: AEMO.", "0.0000"),
    ("tgp", "Petrol TGP", "Petrol — wholesale terminal gate prices", "Cents per litre, GST "
     "inclusive. Seven capitals plus a volume-weighted national average, business days only. "
     "Source: Australian Institute of Petroleum.", "0.00"),
    ("retail", "Petrol retail", "Petrol — retail pump prices", "Cents per litre, GST inclusive. "
     "Unweighted average across reporting sites; coverage differs by state — check the _sites "
     "columns.", "0.00"),
]
# Quarters are written as text (2026-Q1) - Excel has no quarter date format, and a
# quarter-start date would display as an ordinary month and read as one.
FREQS = [("daily", "daily", "yyyy-mm-dd"), ("monthly", "monthly", "yyyy-mm"),
         ("quarterly", "quarterly", "@"), ("annual", "annual", "0")]


def read(path):
    with open(path, newline="") as fh:
        r = csv.reader(fh)
        return next(r), list(r)


def key_cell(ws, row, col, raw, freq, fmt):
    if freq == "annual":
        c = ws.cell(row=row, column=col, value=int(raw))
    elif freq == "quarterly":
        c = ws.cell(row=row, column=col, value=raw)
    elif freq == "monthly":
        y, m = raw.split("-")
        c = ws.cell(row=row, column=col, value=datetime.date(int(y), int(m), 1))
    else:
        c = ws.cell(row=row, column=col, value=datetime.date.fromisoformat(raw))
    c.number_format = fmt
    return c


def add_sheet(wb, name, title, subtitle, header, rows, numfmt, freq, keyfmt):
    ws = wb.create_sheet(name)
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A2"] = subtitle
    ws["A2"].font = SUB_FONT
    hr = 4
    for j, h in enumerate(header, 1):
        c = ws.cell(row=hr, column=j, value=h)
        c.fill, c.font = HDR_FILL, HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(bottom=THIN)
    for i, row in enumerate(rows, hr + 1):
        for j, v in enumerate(row, 1):
            if j == 1:
                key_cell(ws, i, j, v, freq, keyfmt)
            else:
                c = ws.cell(row=i, column=j, value=float(v) if v not in ("", None) else None)
                c.number_format = "0" if header[j - 1].endswith(COUNT_SUFFIXES) else numfmt
    ws.freeze_panes = ws.cell(row=hr + 1, column=2)
    ws.auto_filter.ref = f"A{hr}:{get_column_letter(len(header))}{hr + len(rows)}"
    ws.column_dimensions["A"].width = 12
    for j in range(2, len(header) + 1):
        ws.column_dimensions[get_column_letter(j)].width = max(10, min(16, len(header[j - 1]) * 0.9))
    ws.row_dimensions[hr].height = 42


wb = openpyxl.Workbook()
wb.remove(wb.active)
coverage = []

for stem, label, title, subtitle, numfmt in GROUPS:
    if stem == "retail" and not RETAIL:
        coverage.append((label, "NOT INCLUDED — built with the no-retail option"))
        continue
    spans = []
    for freq, suffix, keyfmt in FREQS:
        path = os.path.join(WORK, f"{stem}_{suffix}.csv")
        hdr, rows = read(path)
        add_sheet(wb, f"{label} {freq}", f"{title} ({freq})", subtitle,
                  hdr, rows, numfmt, freq, keyfmt)
        spans.append(f"{len(rows):,} {freq}")
        if freq == "daily":
            span_txt = f"{rows[0][0]} to {rows[-1][0]}"
    coverage.append((label, f"{span_txt} — " + ", ".join(spans) + " rows"))

write_notes(wb, datetime.date.today().isoformat(), coverage, retail=RETAIL)
write_charts(wb)
wb.save(OUT)
print(f"wrote {OUT} ({len(wb.sheetnames)} tabs)", flush=True)
