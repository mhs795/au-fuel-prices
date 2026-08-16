"""Build daily retail pump price averages for NSW and Queensland.

Both states publish price-CHANGE events, not daily snapshots, so a daily average
requires carrying each site's last posted price forward until it changes. Sites that
have not reported for STALE_DAYS are dropped so closed sites stop dragging the mean.

Monthly aggregates are cached. Because the forward-fill carries across month
boundaries, each month also caches its end-of-month price state, so a rerun only needs
the previous month's state plus the months that can still change.
"""
import calendar, csv, io, json, os, re, sys, urllib.request
from collections import defaultdict
from datetime import datetime, date, timedelta

UA = "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/128.0 Safari/537.36"
OUT = sys.argv[1] if len(sys.argv) > 1 else "state_daily.csv"
CACHE = os.path.join(os.path.dirname(os.path.abspath(OUT)), "state_cache")
STALE_DAYS = 30
FIELDS = ["date", "state", "product", "mean_price_cpl", "n_sites"]

# Rows the readers could not use, counted per month so a source format change shows up
# as a loud number rather than as prices quietly carried forward over the gap.
DROPS = defaultdict(int)

NSW_PKG = "https://data.nsw.gov.au/data/api/3/action/package_show?id=a97a46fc-2bdd-4b90-ac7f-0cb1e8d7ac3b"
QLD_SEARCH = "https://www.data.qld.gov.au/api/3/action/package_search?q=%22Fuel+price+reporting%22&rows=30"
MONTHS = {m.lower(): i for i, m in enumerate(calendar.month_name) if m}
MONTHS.update({m.lower(): i for i, m in enumerate(calendar.month_abbr) if m})

# Source fuel labels -> the common labels used in the workbook
NSW_FUEL = {"U91": "ULP", "E10": "E10", "P95": "PULP95", "P98": "PULP98",
            "DL": "Diesel", "PDL": "Diesel_premium", "LPG": "LPG"}


def norm_key(k):
    """Fold a source column name to a stable lookup key.

    Queensland has published the same column as both Fuel_Type and 'Fuel Type', and
    SiteId as both SiteId and 'Site Id'. Looking the raw name up directly meant a file
    with spaces matched nothing and every row in it was silently skipped.
    """
    return re.sub(r"[^a-z0-9]+", "_", (k or "").strip().lower()).strip("_")


# The NSW archive is not consistent: most months are 24-hour, but the 2019-2023 xlsx
# files are 12-hour with an AM/PM suffix ("1/04/2019 12:08:43 AM"). Uppercased before
# matching so %p takes a lowercase "am" too.
TS_FORMATS = (
    "%d/%m/%Y %I:%M:%S %p", "%d/%m/%Y %I:%M %p",
    "%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y",
    "%Y-%m-%d %I:%M:%S %p", "%Y-%m-%d %I:%M %p",
    "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d",
)


def parse_dt(s):
    """Parse a source timestamp to a naive datetime, or None.

    Sources vary across the archive: ISO separated by a space or by a T, and d/m/Y with
    the day either zero-padded or not. Parsing the whole string rather than slicing a
    fixed width is what makes unpadded days work - an earlier s[:10] slice left a
    trailing space on every 1st-9th of the month and dropped all of them.
    """
    s = (s or "").strip()
    if not s:
        return None
    try:
        return datetime.fromisoformat(s.replace("T", " "))
    except ValueError:
        pass
    for fmt in TS_FORMATS:
        try:
            return datetime.strptime(s.upper(), fmt)
        except ValueError:
            continue
    return None


def qld_fuel(name):
    n = (name or "").strip().lower()
    if "lpg" in n: return "LPG"
    if "diesel" in n:
        return "Diesel_premium" if ("premium" in n or "pdl" in n) else "Diesel"
    if "e10" in n: return "E10"
    if "98" in n: return "PULP98"
    if "95" in n: return "PULP95"
    if "unleaded" in n or n == "ulp": return "ULP"
    return None      # E85, OPAL and anything else are deliberately excluded


def fetch(url, tries=4):
    for a in range(tries):
        try:
            req = urllib.request.Request(url, headers={"User-Agent": UA, "Accept": "*/*"})
            return urllib.request.urlopen(req, timeout=600).read()
        except Exception:
            if a == tries - 1:
                raise
            import time; time.sleep(5 * (a + 1))


def parse_month(name):
    m = re.search(r"(" + "|".join(sorted(MONTHS, key=len, reverse=True)) + r")\w*\s+(\d{4})",
                  (name or "").lower())
    if m:
        return int(m.group(2)), MONTHS[m.group(1)]
    m = re.search(r"(\d{4})[-_](\d{2})", name or "")
    if m and 1 <= int(m.group(2)) <= 12:
        return int(m.group(1)), int(m.group(2))
    return None


def discover():
    """-> {(state, year, month): url}"""
    out = {}
    pkg = json.loads(fetch(NSW_PKG))["result"]
    for r in pkg["resources"]:
        fmt = (r.get("format") or "").lower()
        if not (fmt.startswith("xlsx") or fmt.startswith("excel") or fmt == "csv"):
            continue
        ym = parse_month(r.get("name"))
        if ym:
            out[("NSW", ym[0], ym[1])] = r["url"]
    for p in json.loads(fetch(QLD_SEARCH))["result"]["results"]:
        for r in p.get("resources", []):
            if (r.get("format") or "").upper() != "CSV":
                continue
            ym = parse_month(r.get("name"))
            if ym:
                out[("QLD", ym[0], ym[1])] = r["url"]
    return out


def _nsw_rows(raw):
    """NSW publishes some months as xlsx and some as plain CSV; both share columns."""
    if raw[:2] == b"PK":
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(values_only=True):
            yield row
        wb.close()
    else:
        for row in csv.reader(io.StringIO(raw.decode("utf-8-sig", errors="replace"))):
            yield row


def read_nsw(raw):
    rows = _nsw_rows(raw)
    # some months put a title row above the header, so find the real header row
    hdr = None
    for _ in range(10):
        try:
            cand = [str(h or "").strip() for h in next(rows)]
        except StopIteration:
            break
        if "ServiceStationName" in cand and "Price" in cand:
            hdr = cand
            break
    if hdr is None:
        raise ValueError("no header row found in first 10 rows")
    ix = {h: i for i, h in enumerate(hdr)}
    # the column has been called both FuelCode and FuelType across the archive
    fuel_col = "FuelCode" if "FuelCode" in ix else ("FuelType" if "FuelType" in ix else None)
    need = ("ServiceStationName", "Address", "PriceUpdatedDate", "Price")
    if fuel_col is None or not all(k in ix for k in need):
        raise ValueError(f"unexpected NSW columns: {hdr}")
    for r in rows:
        try:
            price = float(r[ix["Price"]])
        except (TypeError, ValueError):
            continue
        if not (20 <= price <= 600):
            continue
        rawfuel = str(r[ix[fuel_col]] or "").strip()
        # codes (U91/DL/...) in most files, plain names in some of the older ones
        fuel = NSW_FUEL.get(rawfuel.upper()) or qld_fuel(rawfuel)
        if not fuel:
            continue
        ts = r[ix["PriceUpdatedDate"]]
        if hasattr(ts, "date"):
            d = ts.date()                       # NSW timestamps are already local time
        else:
            dt = parse_dt(str(ts))
            if dt is None:
                DROPS["NSW bad timestamp"] += 1
                continue
            d = dt.date()
        site = f"{r[ix['ServiceStationName']]}|{r[ix['Address']]}"
        yield d, site, fuel, price


def read_qld(raw):
    txt = raw.decode("utf-8-sig", errors="replace")
    rdr = csv.DictReader(io.StringIO(txt))
    if rdr.fieldnames:
        have = {norm_key(k) for k in rdr.fieldnames}
        missing = {"fuel_type", "price", "transactiondateutc"} - have
        if missing:
            raise ValueError(f"unexpected QLD columns, missing {sorted(missing)}: "
                             f"{rdr.fieldnames}")
    for raw_row in rdr:
        r = {norm_key(k): v for k, v in raw_row.items()}
        fuel = qld_fuel(r.get("fuel_type"))
        if not fuel:
            continue                              # E85, OPAL and friends, deliberately
        try:
            price = float(r["price"]) / 10.0       # published in tenths of a cent
        except (TypeError, ValueError, KeyError):
            DROPS["QLD bad price"] += 1
            continue
        if not (20 <= price <= 600):
            DROPS["QLD price out of range"] += 1
            continue
        dt = parse_dt(r.get("transactiondateutc"))
        if dt is None:
            DROPS["QLD bad timestamp"] += 1
            continue
        d = (dt + timedelta(hours=10)).date()      # UTC -> AEST; QLD has no DST
        yield d, str(r.get("siteid") or r.get("site_name")), fuel, price


def month_days(y, m):
    n = calendar.monthrange(y, m)[1]
    return [date(y, m, d) for d in range(1, n + 1)]


def process_month(state, y, m, url, carry):
    """carry: {site|fuel: [iso_date, price]} entering the month. Returns (rows, carry_out)."""
    raw = fetch(url)
    reader = read_nsw if state == "NSW" else read_qld
    DROPS.clear()
    events = defaultdict(list)          # day -> list of (key, fuel, price, day)
    n_events = 0
    for d, site, fuel, price in reader(raw):
        events[d].append((f"{site}\x1f{fuel}", fuel, price, d))
        n_events += 1
    # A month that parses to nothing is a source format change, not an empty month. Left
    # unchecked it is invisible: the forward-fill just carries the previous month's
    # prices over the hole and the site count decays as sites go stale.
    if n_events == 0:
        raise ValueError(f"{state} {y}-{m:02d} parsed to 0 usable rows from {len(raw)} "
                         f"bytes (drops: {dict(DROPS) or 'none'})")
    if DROPS:
        total = n_events + sum(DROPS.values())
        share = sum(DROPS.values()) / total
        note = f"  {state} {y}-{m:02d}: dropped {dict(DROPS)} of {total} rows"
        if share > 0.05:
            raise ValueError(note + f" - {share:.1%} is too many to be incidental")
        print(note, flush=True)

    cur = {k: (date.fromisoformat(v[0]), v[1]) for k, v in carry.items()}
    rows = []
    for day in month_days(y, m):
        for key, fuel, price, ed in events.get(day, []):
            cur[key] = (ed, price)
        cutoff = day - timedelta(days=STALE_DAYS)
        acc = defaultdict(lambda: [0.0, 0])
        for key, (upd, price) in cur.items():
            if upd < cutoff or upd > day:
                continue
            a = acc[key.split("\x1f")[1]]
            a[0] += price
            a[1] += 1
        for fuel, (s, c) in sorted(acc.items()):
            if c >= 5:                  # too few sites reporting to be a meaningful average
                rows.append([day.isoformat(), state, fuel, round(s / c, 3), c])
    # drop anything already stale at month end so the carried state cannot grow forever
    end = month_days(y, m)[-1]
    carry_out = {k: [u.isoformat(), p] for k, (u, p) in cur.items()
                 if u >= end - timedelta(days=STALE_DAYS)}
    return rows, carry_out


def main():
    os.makedirs(CACHE, exist_ok=True)
    found = discover()
    keys = sorted(found)
    print(f"{len(keys)} state-months discovered "
          f"(NSW {sum(1 for k in keys if k[0]=='NSW')}, QLD {sum(1 for k in keys if k[0]=='QLD')})",
          flush=True)

    today = date.today()
    live = {(today.year, today.month)}
    pm = date(today.year, today.month, 1) - timedelta(days=1)
    live.add((pm.year, pm.month))

    fetched = reused = 0
    failures = []
    for state in ("NSW", "QLD"):
        months = [k for k in keys if k[0] == state]
        carry = {}
        for (_, y, m) in months:
            agg_p = os.path.join(CACHE, f"{state}-{y}-{m:02d}.csv")
            st_p = os.path.join(CACHE, f"{state}-{y}-{m:02d}.state.json")
            if (y, m) not in live and os.path.exists(agg_p) and os.path.exists(st_p):
                carry = json.load(open(st_p))
                reused += 1
                continue
            try:
                rows, carry = process_month(state, y, m, found[(state, y, m)], carry)
            except Exception as e:
                # Do NOT cache a failed month and do NOT carry its state forward: both
                # would make the failure sticky across reruns and silently fabricate
                # prices for the days it should have covered.
                print(f"  !! {state} {y}-{m:02d} failed: {type(e).__name__}: {e}", flush=True)
                failures.append(f"{state} {y}-{m:02d}: {type(e).__name__}: {e}")
                carry = {}
                continue
            with open(agg_p + ".tmp", "w", newline="") as fh:
                w = csv.writer(fh); w.writerow(FIELDS); w.writerows(rows)
            os.replace(agg_p + ".tmp", agg_p)
            json.dump(carry, open(st_p + ".tmp", "w"))
            os.replace(st_p + ".tmp", st_p)
            fetched += 1
            print(f"  {state} {y}-{m:02d}: {len(rows)} daily rows, {len(carry)} sites carried",
                  flush=True)

    print(f"{fetched} month(s) processed, {reused} reused from cache", flush=True)
    out = []
    for name in sorted(os.listdir(CACHE)):
        if name.endswith(".csv"):
            with open(os.path.join(CACHE, name), newline="") as fh:
                out.extend(list(csv.DictReader(fh)))
    # Bail out BEFORE writing, so a failed run leaves the previous complete file in place
    # rather than replacing it with a short one.
    if failures:
        raise SystemExit("retail source failures - refusing to write a short series:\n  "
                         + "\n  ".join(failures))
    out.sort(key=lambda r: (r["date"], r["state"], r["product"]))
    with open(OUT, "w", newline="") as fh:
        w = csv.DictWriter(fh, FIELDS); w.writeheader(); w.writerows(out)
    print("wrote", OUT, len(out), "rows", flush=True)


if __name__ == "__main__":
    main()
