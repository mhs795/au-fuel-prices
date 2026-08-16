"""Content and writer for the 'Summary charts' tab.

Every chart points at cells on the data tabs - nothing is copied or recomputed here, so
a chart can never drift from the numbers it claims to show. Rebuilding the workbook
rebuilds the charts from the same source rows.

Colours are a fixed categorical order, assigned per series and never cycled: a series
keeps its colour across every chart it appears in, so NSW is the same navy on the daily,
monthly, quarterly and annual electricity charts.

The palette is taken from the Commonwealth Budget papers - sampled directly out of the
vector graphics in Budget Paper No. 1 2026-27, Statements 2 and 7, not eyeballed - so
these charts sit alongside Budget material without looking foreign. Gridline, axis and
heading greys come from the same source.

The honest cost of that choice: the Budget palette is built for charts carrying two or
three series, and these carry five or six, so it is being stretched. Measured separation
falls from the previous palette's worst pair of dE 39.6 (normal vision) and 28.6
(protanopia) to dE 18.9 and 15.5 here, the weak pair being navy 002A54 against indigo
212F73. That is still readable, but it is roughly half the headroom. It is acceptable
only because identity never rests on colour alone - every chart carries a legend, series
are ordered consistently, and the underlying table is one tab away. If these charts ever
need to work for someone reading them at a distance or with a colour-vision deficiency,
switch PALETTE back to the categorical set kept in PALETTE_CVD below.
"""
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.text import RichText
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.text import (CharacterProperties, Paragraph, ParagraphProperties,
                                   RichTextProperties)
from openpyxl.styles import Font

# Commonwealth Budget palette, sampled from the vector graphics of Budget Paper No. 1
# 2026-27 (Statements 2 and 7). Ordered for maximum separation: navy, bright blue, green,
# teal, indigo - chosen by measuring every 5-subset rather than by eye.
PALETTE = ["002A54", "417AE4", "5B9866", "338E8B", "212F73", "626A77", "003E18", "293F5B"]
# The previous palette, kept because it separates far better when more than three series
# share a chart - see the note in the module docstring. Swap PALETTE = PALETTE_CVD to use it.
PALETTE_CVD = ["2A78D6", "EB6834", "1BAF7A", "EDA100", "E87BA4", "008300", "4A3AA7", "E34948"]
GRID = "E1E3E6"      # Budget gridline grey
AXIS = "626A77"      # Budget axis and axis-label grey
INK = "002A54"       # Budget heading navy
HDR_ROW = 4          # data tabs put their header on row 4
FIRST_DATA_ROW = 5
LINE_W = 19050       # ~1.5pt in EMU

TITLE_FONT = Font(bold=True, size=13, color=INK)
SECTION_FONT = Font(bold=True, size=12, color=INK)
SUBSECTION_FONT = Font(bold=True, size=10, color=AXIS)
SUB_FONT = Font(italic=True, size=9, color=AXIS)

# The petrol charts are pinned to a common scale so that wholesale and retail, and each
# frequency, can be read against one another without rescaling by eye. The band is set
# wide enough to contain every charted petrol value - the 2020 TGP trough at 80.8 c/L and
# the April 2026 peak at 325.9 - so nothing is drawn off the edge of a plot. Widen it if a
# future refresh moves outside; the assertion in check_ylim() will say so.
# Electricity and gas stay auto-scaled - their spikes are orders of magnitude, not
# percentages, and a fixed range would either flatten them or clip them off.
PETROL_YLIM = (75, 350, 50)   # min, max, gridline spacing

# (data tab stem, chart title, y-axis title, [(column header, legend name), ...], y-range)
SERIES = [
    ("Electricity", "Electricity spot price (DWA)", "$/MWh",
     [("NSW1_price_dwa", "NSW"), ("QLD1_price_dwa", "QLD"), ("SA1_price_dwa", "SA"),
      ("TAS1_price_dwa", "TAS"), ("VIC1_price_dwa", "VIC")], None),
    ("Gas", "Gas wholesale price", "$/GJ",
     [("VIC_DWGM_6am", "VIC DWGM"), ("SYD_exante", "Sydney STTM"),
      ("ADL_exante", "Adelaide STTM"), ("BRI_exante", "Brisbane STTM"),
      # Wallumbilla sits on the same axis as the markets it competes with, which is
      # the point of putting it here: it is an upstream traded price against city-gate
      # clearing prices, and the comparison is the information. Its slot continues
      # after the four above so no colour is reused.
      # The hub's SEQ location is deliberately NOT charted. It trades on 77% of days,
      # correlates 0.91 with Brisbane STTM sitting about 40c below it, and so mostly
      # redraws the Brisbane line with holes in it. The columns stay on the data tabs.
      ("GSH_WAL_price", "Wallumbilla GSH", 4)], None),
    ("Petrol TGP", "Petrol wholesale TGP, national", "cents/litre",
     [("TGP_petrol_national", "Petrol"), ("TGP_diesel_national", "Diesel")], PETROL_YLIM),
    ("Petrol retail", "Retail ULP pump price", "cents/litre",
     [("NSW_ULP", "NSW"), ("QLD_ULP", "QLD"), ("WA_ULP", "WA")], PETROL_YLIM),
]

# Row selectors. None is the whole tab; ("window", n) is the trailing n years measured
# from that series' own last date; ("year", y) is calendar year y. The fourth item is an
# axis override for that chart alone; without one the series' own y-range applies.
STANDARD_VARIANTS = [
    ("daily", "daily, last 2 years", ("window", 2), None),
    ("daily", "daily, last 5 years", ("window", 5), None),
    ("monthly", "monthly", None, None),
    ("quarterly", "quarterly", None, None),
    ("annual", "annual", None, None),
]
# Petrol also gets a chart per recent calendar year, each on a tighter axis of its own so
# a single year's movement is legible rather than flattened into the full 75-350 band.
# 2026 has to reach 350 rather than a tighter ceiling: TGP diesel peaked at 325.9 c/L on
# 9 April 2026, and anything lower crops that spike off the top of the plot.
YLIM_2026 = (125, 350, 25)
YLIM_2025 = (125, 200, 25)
PETROL_VARIANTS = STANDARD_VARIANTS[:2] + [
    ("daily", "daily, 2026", ("year", 2026), YLIM_2026),
    ("daily", "daily, 2025", ("year", 2025), YLIM_2025),
] + STANDARD_VARIANTS[2:]

# Daily retail petrol, one state per chart, for a single calendar year. The all-states
# chart stacks NSW, QLD and WA on one plot, which is exactly where a day-to-day reading
# stops being possible: three discount cycles running out of phase read as noise rather
# than as three cycles. Splitting them apart is the only way to see a single state's own
# week. Wholesale TGP is deliberately not split this way - the capitals sit within a
# couple of cents of each other at the terminal gate, so the national line already tells
# that story and seven near-identical charts would only pad the tab.
# The states share one axis so the charts stay comparable side by side, tighter than the
# 75-350 band because one year of one market moves within a few tens of cents. The band is
# verified against the data on every build.
STATE_YEAR = 2025
RETAIL_STATE_YLIM = (155, 200, 5)    # 2025 state ULP spans 162.4-195.2 c/L

# {data tab stem: (chart title, y-range, [(label, [(header, legend, palette slot)])])}
# The palette slot is given explicitly because a series has to keep the colour it carries
# on the all-states chart: NSW retail is the same blue whether it is drawn beside QLD and
# WA or on its own. Left implicit it would restart at slot 0 on every single-state chart.
PER_STATE = {
    "Petrol retail": ("Retail ULP pump price", RETAIL_STATE_YLIM, [
        (state, [(f"{state}_ULP", f"{state} ULP", k)])
        for k, state in enumerate(["NSW", "QLD", "WA"])]),
}

# Heading and standfirst for each block of charts on the summary tab, keyed by data tab
# stem. A heading is only written if the block below it actually has charts, so a
# no-retail build does not leave a heading standing over nothing.
SECTIONS = {
    "Electricity": ("Electricity",
                    "NEM regional spot prices, demand-weighted. Monthly and coarser only "
                    "— daily spot is too spiky to read at this size."),
    "Gas": ("Gas",
            "Victorian DWGM schedule prices, the three STTM hub ex-ante prices, and the "
            "Gas Supply Hub at Wallumbilla, $/GJ. The hub's data starts in 2026, so it "
            "is a short line at the right-hand end of the longer charts and absent from "
            "the annual one. The hub's SEQ prices are on the data tabs, not charted."),
    "Petrol TGP": ("Petrol — wholesale (terminal gate)",
                   "AIP terminal gate prices, cents per litre, business days only. The "
                   "cost side of the pump price, before retail margin and the discount "
                   "cycle."),
    "Petrol retail": ("Petrol — retail (pump)",
                      "Average pump price across reporting sites, cents per litre. "
                      "Coverage differs by state — check the _sites columns on the data "
                      f"tab. The single-state charts cover {STATE_YEAR}."),
}

# Series that get no daily chart on the summary tab. Electricity is here: half-hourly
# spot averaged to a day is too spiky to read at chart size, and the monthly, quarterly
# and annual charts carry the same story. The "Electricity daily" data tab is unaffected.
NO_DAILY_CHARTS = {"Electricity"}

# One electricity/gas pair per state, each on its own two-scale chart. Tasmania has no
# gas hub - it is not in the STTM and has no DWGM - so it has no chart here; TAS spot
# prices are on the electricity tabs. Each pair starts where its gas series starts,
# which differs by hub.
STATE_PAIRS = [
    ("New South Wales", "NSW1_price_dwa", "NSW spot", "SYD_exante", "Sydney STTM"),
    ("Victoria", "VIC1_price_dwa", "VIC spot", "VIC_DWGM_6am", "VIC DWGM"),
    ("Queensland", "QLD1_price_dwa", "QLD spot", "BRI_exante", "Brisbane STTM"),
    ("South Australia", "SA1_price_dwa", "SA spot", "ADL_exante", "Adelaide STTM"),
]


def col_index(ws, header):
    """1-based column of a header on a data tab, or None if the tab lacks it."""
    for j in range(1, ws.max_column + 1):
        if ws.cell(row=HDR_ROW, column=j).value == header:
            return j
    return None


def search_rows(ws, ok):
    """First and last row whose date column A satisfies ok(), or None if none do."""
    rows = [i for i in range(FIRST_DATA_ROW, ws.max_row + 1)
            if ok(ws.cell(row=i, column=1).value)]
    return (rows[0], rows[-1]) if rows else None


def select_rows(ws, selector):
    """Resolve a variant's row selector to (first_row, last_row)."""
    if selector is None:
        return FIRST_DATA_ROW, ws.max_row
    kind, n = selector
    if kind == "window":
        cutoff = ws.cell(row=ws.max_row, column=1).value.toordinal() - round(365.25 * n)
        return search_rows(ws, lambda d: d.toordinal() >= cutoff)
    if kind == "year":
        return search_rows(ws, lambda d: d.year == n)
    raise ValueError(f"unknown row selector {selector!r}")


def year_label(ws, first_row, last_row, year):
    """Say so when a calendar-year chart is only part of a year."""
    last = ws.cell(row=last_row, column=1).value
    if (last.month, last.day) == (12, 31):
        return f"daily, {year}"
    return f"daily, {year} (to {last.strftime('%-d %b')})"


def row_of(ws, key):
    """Row holding a given period key, matching on the tab's own column A values."""
    for i in range(FIRST_DATA_ROW, ws.max_row + 1):
        if ws.cell(row=i, column=1).value == key:
            return i
    return None


def style_series(s, colour):
    s.graphicalProperties = GraphicalProperties(
        ln=LineProperties(solidFill=colour, w=LINE_W))
    s.smooth = False
    s.marker.symbol = "none"


def recede_axes(chart, n_points):
    gl = ChartLines()
    gl.spPr = GraphicalProperties(ln=LineProperties(solidFill=GRID, w=9525))
    chart.y_axis.majorGridlines = gl
    chart.x_axis.majorGridlines = None
    for ax in (chart.x_axis, chart.y_axis):
        ax.delete = False
        ax.spPr = GraphicalProperties(ln=LineProperties(solidFill=AXIS, w=9525))
    # openpyxl defaults BOTH axes to axPos "l", which is invalid - the category axis
    # belongs at the bottom. Excel on the desktop honours what the file says and lays the
    # chart out from it, so the category labels end up fighting the value axis and its
    # title. Web and phone viewers re-layout and hide the problem, which is why this only
    # showed up in Office. Setting each axis to its real position fixes it.
    chart.x_axis.axPos = "b"
    chart.y_axis.axPos = "l"
    # Be explicit about label placement too, rather than leaving Excel to infer it.
    chart.x_axis.tickLblPos = "nextTo"
    chart.y_axis.tickLblPos = "nextTo"
    # Real tick marks: without them a slanted label has nothing to point at, so a reader
    # cannot tell which period a label belongs to once labels are being skipped.
    chart.x_axis.majorTickMark = "out"
    chart.y_axis.majorTickMark = "out"
    skip = max(1, n_points // 8)
    chart.x_axis.tickLblSkip = skip
    chart.x_axis.tickMarkSkip = skip
    # Period labels are wide enough to collide when laid flat, so lay them on the slant.
    small = CharacterProperties(sz=800)
    chart.x_axis.txPr = RichText(
        bodyPr=RichTextProperties(rot=-2700000, vert="horz"),
        p=[Paragraph(pPr=ParagraphProperties(defRPr=small), endParaRPr=small)])



def check_ylim(ws, cols, first_row, last_row, ylim, chart_title, deliberate):
    """A fixed axis hides anything outside it, so never let that pass unremarked.

    The shared petrol band is meant to contain everything, so anything outside it is a
    mistake and stops the build. An axis chosen for one chart is an editorial decision -
    it is allowed to crop, but every build says what it crops, so the cost stays visible.
    """
    lo, hi = ylim[0], ylim[1]
    outside = []
    for col in cols:
        header = col[0]
        j = col_index(ws, header)
        if j is None:
            continue
        for i in range(first_row, last_row + 1):
            v = ws.cell(row=i, column=j).value
            if v is not None and not lo <= v <= hi:
                outside.append((ws.cell(row=i, column=1).value, header, v))
    if not outside:
        return
    if not deliberate:
        d, header, v = outside[0]
        raise SystemExit(
            f"ERROR: {ws.title} {header} = {v} on {d} falls outside the fixed axis range "
            f"{lo}-{hi} of '{chart_title}', so the chart would draw it off the plot. "
            f"Widen that chart's axis constant in build_charts.py.")
    days = sorted({d for d, _, _ in outside})
    worst = max(outside, key=lambda t: abs(t[2] - (hi if t[2] > hi else lo)))
    print(f"  note: '{chart_title}' axis {lo}-{hi} crops {len(days)} day(s), "
          f"{days[0]:%Y-%m-%d} to {days[-1]:%Y-%m-%d}; furthest is {worst[1]} "
          f"{worst[2]} on {worst[0]:%Y-%m-%d}", flush=True)


def line_chart(ws, title, y_title, cols, first_row, last_row, ylim=None,
               width=15.5, height=8.2):
    chart = LineChart()
    chart.title = title
    chart.y_axis.title = y_title
    chart.height, chart.width = height, width
    if ylim:
        chart.y_axis.scaling.min, chart.y_axis.scaling.max, chart.y_axis.majorUnit = ylim
    for k, col in enumerate(cols):
        header, name = col[0], col[1]
        slot = col[2] if len(col) > 2 else k     # explicit slot keeps a series' colour
        j = col_index(ws, header)
        if j is None:
            continue
        # A series with one point in range draws nothing - the markers are off - but
        # would still claim a legend entry, so the chart would name a line that is not
        # there. Short series drop out and reappear on their own as history accumulates:
        # the Gas Supply Hub starts in 2026, so it is on the daily and monthly gas charts
        # from the start and joins the annual one once it has two years.
        if populated(ws, j, first_row, last_row) < 2:
            continue
        s = Series(Reference(ws, min_col=j, min_row=first_row, max_row=last_row),
                   title=name)
        style_series(s, PALETTE[slot % len(PALETTE)])
        chart.append(s)
    chart.set_categories(Reference(ws, min_col=1, min_row=first_row, max_row=last_row))
    recede_axes(chart, last_row - first_row + 1)
    if len(chart.series) > 1:
        chart.legend.position = "b"
        chart.legend.overlay = False
    else:
        chart.legend = None
    return chart


def populated(ws, col, first_row, last_row):
    """How many rows in range actually carry a value in that column."""
    return sum(ws.cell(row=i, column=col).value is not None
               for i in range(first_row, last_row + 1))


def first_populated_row(ws, col):
    """First row where a column actually has a number - hubs start at different dates."""
    for i in range(FIRST_DATA_ROW, ws.max_row + 1):
        if ws.cell(row=i, column=col).value is not None:
            return i
    return None


def combined_chart(wb, state, elec_col, elec_name, gas_col, gas_name):
    """One state's electricity and gas, quarterly, with gas on a secondary axis.

    Both series are from the same state, so the two lines describe one market rather than
    two unrelated ones. Each chart starts at the first quarter its gas hub published -
    2007-Q1 for the Victorian DWGM, 2010-Q3 for Sydney and Adelaide, 2011-Q4 for Brisbane
    - because charting from 1998 would leave most of the plot with only one line on it.

    The two prices are on genuinely different scales ($/MWh against $/GJ), so read the
    shapes and their timing against each other, not the crossings - where the lines sit
    relative to one another is an artefact of the two axis ranges, not a fact about the
    market.
    """
    elec, gas = wb["Electricity quarterly"], wb["Gas quarterly"]
    g0 = first_populated_row(gas, col_index(gas, gas_col))
    g1 = gas.max_row
    e0, e1 = row_of(elec, gas.cell(row=g0, column=1).value), elec.max_row
    if e1 - e0 != g1 - g0:
        raise SystemExit(f"ERROR: {state} quarterly rows do not line up between the "
                         f"electricity and gas tabs ({e1 - e0 + 1} vs {g1 - g0 + 1}); "
                         f"the two lines would be drawn against the wrong quarters.")

    c1 = LineChart()
    c1.title = f"Electricity and gas — quarterly, {state} (two scales)"
    c1.y_axis.title = "Electricity $/MWh"
    c1.y_axis.axId = 100
    s = Series(Reference(elec, min_col=col_index(elec, elec_col), min_row=e0, max_row=e1),
               title=f"{elec_name} ($/MWh, left)")
    style_series(s, PALETTE[0])
    c1.append(s)
    c1.set_categories(Reference(elec, min_col=1, min_row=e0, max_row=e1))
    recede_axes(c1, e1 - e0 + 1)
    c1.height, c1.width = 8.2, 15.5

    c2 = LineChart()
    c2.y_axis.axId = 200
    c2.y_axis.title = "Gas $/GJ"
    c2.y_axis.majorGridlines = None
    c2.y_axis.delete = False
    c2.y_axis.spPr = GraphicalProperties(ln=LineProperties(solidFill=AXIS, w=9525))
    s = Series(Reference(gas, min_col=col_index(gas, gas_col), min_row=g0, max_row=g1),
               title=f"{gas_name} ($/GJ, right)")
    # Green, not the adjacent blue: on the Budget palette slots 0 and 1 are both blues,
    # and two blues on a two-scale chart is exactly where a reader stops being able to
    # tell which line belongs to which axis.
    style_series(s, PALETTE[2])
    c2.append(s)

    # Draws the gas axis on the right. This must be set on the secondary axis: setting
    # it on the primary instead moves the electricity axis across, leaving each title on
    # the opposite side from its own scale.
    c2.y_axis.crosses = "max"
    # Same axPos correction as recede_axes makes, for the axes this chart owns itself: the
    # gas scale genuinely is on the right, and c2's own category axis is the hidden
    # duplicate that carries the pairing, so it is deleted rather than drawn twice.
    c2.y_axis.axPos = "r"
    c2.y_axis.tickLblPos = "nextTo"
    c2.y_axis.majorTickMark = "out"
    c2.x_axis.axPos = "b"
    c2.x_axis.delete = True
    c1 += c2
    c1.legend.position = "b"
    c1.legend.overlay = False
    return c1


class Grid:
    """A two-column run of charts, top to bottom, broken up by section headings.

    A heading always starts a fresh row, so a section never begins in the right-hand
    column beside the tail of the one before it - which would put a chart above its own
    heading and under someone else's.
    """

    PITCH = 18          # rows a chart occupies, at the 8.2cm chart height
    HEADING = 3         # heading, standfirst, and a blank row of air beneath

    def __init__(self, ws, row=4):
        self.ws, self.row, self.right = ws, row, False

    def _newline(self):
        if self.right:
            self.row += self.PITCH
            self.right = False

    def heading(self, text, standfirst):
        self._newline()
        if self.row > 4:                    # air above every heading but the first
            self.row += 1
        c = self.ws.cell(row=self.row, column=2, value=text)
        c.font = SECTION_FONT
        c = self.ws.cell(row=self.row + 1, column=2, value=standfirst)
        c.font = SUB_FONT
        self.row += self.HEADING

    def subheading(self, text):
        self._newline()
        c = self.ws.cell(row=self.row, column=2, value=text)
        c.font = SUBSECTION_FONT
        self.row += 2

    def add(self, chart):
        self.ws.add_chart(chart, f"{'L' if self.right else 'B'}{self.row}")
        self.right = not self.right
        if not self.right:
            self.row += self.PITCH


def year_charts(wb, stem, y_title):
    """The one-state-per-chart daily charts for STATE_YEAR, or [] if this stem has none."""
    name = f"{stem} daily"
    if stem not in PER_STATE or name not in wb.sheetnames:
        return []
    title, ylim, groups = PER_STATE[stem]
    data = wb[name]
    rows = select_rows(data, ("year", STATE_YEAR))
    if rows is None:                         # a series that does not reach that year
        return []
    first, last = rows
    charts = []
    for label, cols in groups:
        if not any(col_index(data, c[0]) for c in cols):
            continue
        chart_title = f"{title} — {label}, {year_label(data, first, last, STATE_YEAR)}"
        check_ylim(data, cols, first, last, ylim, chart_title, deliberate=False)
        charts.append(line_chart(data, chart_title, y_title, cols, first, last, ylim=ylim))
    return charts


def frequency_charts(wb, stem, title, y_title, cols, ylim):
    """Every charted frequency for one series, in the order the variants are declared."""
    variants = PETROL_VARIANTS if ylim else STANDARD_VARIANTS
    if stem in NO_DAILY_CHARTS:
        variants = [v for v in variants if v[0] != "daily"]
    charts = []
    for freq, label, selector, ylim_override in variants:
        name = f"{stem} {freq}"
        if name not in wb.sheetnames:
            continue
        data = wb[name]
        rows = select_rows(data, selector)
        if rows is None:                     # a year the series does not reach
            continue
        first, last = rows
        if selector and selector[0] == "year":
            label = year_label(data, first, last, selector[1])
        axis = ylim_override or ylim
        chart_title = f"{title} — {label}"
        if axis:
            check_ylim(data, cols, first, last, axis, chart_title,
                       deliberate=ylim_override is not None)
        charts.append(line_chart(data, chart_title, y_title, cols, first, last, ylim=axis))
    return charts


def write_charts(wb):
    ws = wb.create_sheet("Summary charts", 1)
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Summary charts"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = ("Each series at every frequency it is charted at, drawn live from the "
                "data tabs — the numbers behind any chart are on the matching tab, and "
                "the headings below say which. Daily is shown for gas and petrol over "
                "the last two years and the last five, and for petrol over 2026 and "
                f"{STATE_YEAR} separately, with retail also split one state per chart "
                f"for {STATE_YEAR}; "
                "electricity starts at monthly, because daily spot is too spiky to read "
                "at this size. Monthly, quarterly and annual show full history. A series "
                "keeps the same colour across all of its charts. The all-frequency petrol "
                "charts are pinned to a fixed 75–350 c/L scale so they can be read "
                "against each other; the single-year charts use tighter scales of their "
                "own. The four two-scale charts pair each state's electricity with its "
                "own gas hub.")
    ws["A2"].font = SUB_FONT

    grid = Grid(ws)
    grid.heading("Electricity and gas by state",
                 "One state's electricity and its own gas hub, quarterly, on two scales. "
                 "Read the shapes and their timing against each other, not the crossings.")
    for pair in STATE_PAIRS:
        grid.add(combined_chart(wb, *pair))

    for stem, title, y_title, cols, ylim in SERIES:
        charts = frequency_charts(wb, stem, title, y_title, cols, ylim)
        by_state = year_charts(wb, stem, y_title)
        if not charts and not by_state:      # e.g. retail, in a no-retail build
            continue
        grid.heading(*SECTIONS[stem])
        for chart in charts:
            grid.add(chart)
        if by_state:
            grid.subheading(f"One state per chart — daily, {STATE_YEAR}")
            for chart in by_state:
                grid.add(chart)

    ws.column_dimensions["A"].width = 2
    return ws
