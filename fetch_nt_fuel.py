"""Download MyFuel NT monthly workbooks and aggregate to daily averages.

The Northern Territory is the one mainland jurisdiction whose retail price archive is
openly downloadable in bulk, so it is the only one that can join WA, NSW and QLD on the
retail tabs. SA, TAS and the ACT run real-time schemes whose data is released only to
approved publishers under an API agreement, and Victoria has no scheme at all.

Two things make this simpler than fetch_state_retail.py: MyFuel publishes a *daily
snapshot* of every site rather than price-change events, so there is no forward-fill and
no month-to-month carried state, and every file is final, so a cached month is reused
until its published last_modified changes.

Two things make it messier:
  - A zero is not a price. The workbooks carry a column per product and write 0 where the
    site does not sell it, so zeros must be dropped rather than averaged.
  - The files overlap. Filenames like 'treasury-report-31-oct-to-01-dec-2024.xlsx' run a
    day past each end, so the same site-day appears in two workbooks. Aggregates cannot
    be summed across files without double counting, so where two files cover one day the
    one with more observations wins - it is the file that owns that day.

MyFuel NT stopped publishing after November 2024. The series is therefore closed, not
live; it is kept because seven years of NT history is real published data and nothing
else covers the Territory.
"""
import calendar, csv, io, json, os, re, sys, urllib.request
from collections import defaultdict
from datetime import datetime, date

UA = "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/128.0 Safari/537.36"
OUT = sys.argv[1] if len(sys.argv) > 1 else "nt_daily.csv"
CACHE = os.path.join(os.path.dirname(os.path.abspath(OUT)), "nt_cache")
FIELDS = ["date", "state", "product", "mean_price_cpl", "n_sites"]
SEARCH = "https://data.nt.gov.au/api/3/action/package_search?q=myfuel&rows=100"
MIN_SITES = 5          # the floor WA, NSW and QLD all use
PRICE_LO, PRICE_HI = 20.0, 600.0   # the sanity band used on the NSW and QLD feeds

# MyFuel's column headings -> the common labels shared with WA, NSW and QLD.
# Low Aromatic Fuel, Ethanol 105 (E85) and Bio Diesel 20 are deliberately excluded, as
# E85 and OPAL are on the other feeds.
NT_FUEL = {"Unleaded 91": "ULP", "Ethanol 94 (E10)": "E10", "Premium 95": "PULP95",
           "Premium 98": "PULP98", "Diesel": "Diesel", "LPG": "LPG",
           "Premium Diesel": "Diesel_premium"}
KEEP = {"ULP", "E10", "PULP95", "PULP98", "Diesel", "LPG"}

DROPS = defaultdict(int)


def fetch(url, tries=4):
    for a in range(tries):
        try:
            req = urllib.request.Request(url, headers={"User-Agent": UA, "Accept": "*/*"})
            return urllib.request.urlopen(req, timeout=600).read()
        except Exception:
            if a == tries - 1:
                raise
            import time; time.sleep(5 * (a + 1))


# Most workbooks store FullDate as a real datetime, but the July, October and November
# 2023 files store it as text in 12-hour form ('1/07/2023 12:00:00 AM'). Parsed with
# fromisoformat alone those three months dropped every row.
TS_FORMATS = (
    "%d/%m/%Y %I:%M:%S %p", "%d/%m/%Y %I:%M %p",
    "%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y",
    "%Y-%m-%d %I:%M:%S %p", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d",
)


def parse_dt(s):
    s = (s or "").strip()
    if not s:
        return None
    try:
        return datetime.fromisoformat(s.replace("T", " "))
    except ValueError:
        pass
    for fmt in TS_FORMATS:
        try:
            return datetime.strptime(s.upper(), fmt)
        except ValueError:
            continue
    return None


def rows_of(raw):
    """MyFuel is xlsx throughout, but read a csv too rather than assume."""
    if raw[:2] == b"PK":
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        for row in wb[wb.sheetnames[0]].iter_rows(values_only=True):
            yield row
        wb.close()
    else:
        for row in csv.reader(io.StringIO(raw.decode("utf-8-sig", errors="replace"))):
            yield row


def read_nt(raw):
    """-> (date, site_key, product, price). Finds the header; older files lead with blanks."""
    it = rows_of(raw)
    hdr = None
    for _ in range(20):
        try:
            cand = [str(h or "").strip() for h in next(it)]
        except StopIteration:
            break
        if "FullDate" in cand and "Region Name" in cand:
            hdr = cand
            break
    if hdr is None:
        raise ValueError("no header row found in the first 20 rows")
    ix = {h: i for i, h in enumerate(hdr)}
    fuel_cols = [(i, NT_FUEL[h]) for h, i in ix.items() if h in NT_FUEL]
    if not fuel_cols:
        raise ValueError(f"no recognised fuel columns: {hdr}")
    di, ri, si = ix["FullDate"], ix["Region Name"], ix.get("Suburb", ix["Region Name"])
    lat, lon = ix.get("Lat"), ix.get("Long")

    for r in it:
        v = r[di] if di < len(r) else None
        if hasattr(v, "date"):
            d = v.date()
        else:
            t = str(v or "").strip()
            if not t:
                continue                      # trailing blank rows, common in these files
            dt = parse_dt(t)
            if dt is None:
                DROPS["NT bad date"] += 1
                continue
            d = dt.date()
        region = str(r[ri] or "").strip()
        if not region:
            DROPS["NT no region"] += 1
            continue
        # Lat/Long identify a site far more reliably than the free-text suburb, which is
        # capitalised inconsistently across the archive ('Casuarina' and 'CASUARINA').
        key = (f"{r[lat]},{r[lon]}" if lat is not None and lon is not None
               else f"{r[si]}|{region}")
        for i, prod in fuel_cols:
            if prod not in KEEP or i >= len(r):
                continue
            try:
                price = float(r[i])
            except (TypeError, ValueError):
                continue
            if price <= 0:
                continue                      # 0 means the site does not sell it
            if not (PRICE_LO <= price <= PRICE_HI):
                DROPS["NT price out of range"] += 1
                continue
            yield d, key, prod, price, region


def aggregate(raw):
    """-> {(iso_date, scope, product): [sum, n]} over both scopes."""
    acc = defaultdict(lambda: [0.0, 0])
    seen = set()
    n_rows = 0
    for d, key, prod, price, region in read_nt(raw):
        # One site can appear twice in a day if it posted twice; count it once so the
        # site count stays a site count.
        if (d, key, prod) in seen:
            continue
        seen.add((d, key, prod))
        n_rows += 1
        iso = d.isoformat()
        scopes = ["NT"] + (["Darwin"] if region == "Darwin" else [])
        for scope in scopes:
            e = acc[(iso, scope, prod)]
            e[0] += price
            e[1] += 1
    if n_rows == 0:
        raise ValueError(f"parsed to 0 usable prices (drops: {dict(DROPS) or 'none'})")
    if DROPS:
        total = n_rows + sum(DROPS.values())
        share = sum(DROPS.values()) / total
        note = f"    dropped {dict(DROPS)} of {total}"
        if share > 0.05:
            raise ValueError(note + f" - {share:.1%} is too many to be incidental")
        print(note, flush=True)
    return acc


def discover():
    """-> [(resource_id, name, url, last_modified)] for every readable MyFuel workbook."""
    out = {}
    res = json.loads(fetch(SEARCH))["result"]["results"]
    for p in res:
        for r in p.get("resources", []):
            fmt = (r.get("format") or "").strip().lower()
            url = (r.get("url") or "").strip()
            ok = (fmt.startswith("xlsx") or fmt.startswith("excel") or fmt == "csv"
                  or url.lower().endswith((".xlsx", ".csv")))
            if not ok:
                continue
            stamp = r.get("last_modified") or r.get("created") or ""
            out[r["id"]] = (r["id"], r.get("name") or r["id"], url, stamp)
    return sorted(out.values(), key=lambda t: t[1])


def main():
    os.makedirs(CACHE, exist_ok=True)
    resources = discover()
    if not resources:
        raise SystemExit("NT: no MyFuel resources discovered")
    print(f"{len(resources)} MyFuel NT workbook(s) discovered", flush=True)

    fetched = reused = 0
    failures = []
    for rid, name, url, stamp in resources:
        agg_p = os.path.join(CACHE, f"{rid}.csv")
        meta_p = os.path.join(CACHE, f"{rid}.meta.json")
        if os.path.exists(agg_p) and os.path.exists(meta_p):
            try:
                if json.load(open(meta_p)).get("last_modified") == stamp:
                    reused += 1
                    continue
            except (ValueError, OSError):
                pass
        DROPS.clear()
        try:
            acc = aggregate(fetch(url))
        except Exception as e:
            # Never cache a failed file: a cached failure is a permanent silent hole.
            print(f"  !! {name} failed: {type(e).__name__}: {e}", flush=True)
            failures.append(f"{name}: {type(e).__name__}: {e}")
            continue
        rows = [[iso, scope, prod, round(s / n, 3), n]
                for (iso, scope, prod), (s, n) in sorted(acc.items()) if n >= MIN_SITES]
        with open(agg_p + ".tmp", "w", newline="") as fh:
            w = csv.writer(fh); w.writerow(FIELDS); w.writerows(rows)
        os.replace(agg_p + ".tmp", agg_p)
        json.dump({"last_modified": stamp, "name": name}, open(meta_p + ".tmp", "w"))
        os.replace(meta_p + ".tmp", meta_p)
        fetched += 1
        days = {r[0] for r in rows}
        print(f"  {name}: {len(rows)} rows over {len(days)} days", flush=True)

    print(f"{fetched} workbook(s) processed, {reused} reused from cache", flush=True)
    if failures:
        raise SystemExit("NT source failures - refusing to write a short series:\n  "
                         + "\n  ".join(failures))

    # Union the cached files. Where two workbooks cover the same day - the archive
    # overlaps a day at each end - the one with more observations owns it; adding them
    # would double count the sites they share.
    best = {}
    for fn in sorted(os.listdir(CACHE)):
        if not fn.endswith(".csv"):
            continue
        with open(os.path.join(CACHE, fn), newline="") as fh:
            for r in csv.DictReader(fh):
                k = (r["date"], r["state"], r["product"])
                n = int(r["n_sites"])
                if k not in best or n > int(best[k]["n_sites"]):
                    best[k] = r
    out = sorted(best.values(), key=lambda r: (r["date"], r["state"], r["product"]))
    with open(OUT + ".tmp", "w", newline="") as fh:
        w = csv.DictWriter(fh, FIELDS); w.writeheader(); w.writerows(out)
    os.replace(OUT + ".tmp", OUT)
    days = sorted({r["date"] for r in out})
    print(f"wrote {OUT} rows {len(out)} days {len(days)} {days[0]} -> {days[-1]}", flush=True)


if __name__ == "__main__":
    main()
