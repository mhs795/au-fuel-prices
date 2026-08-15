"""Build a daily east-coast wholesale gas price series ($/GJ) from two AEMO
master workbooks: the Victorian DWGM prices-and-demand file and the STTM
price-and-withdrawals file.

  build_gas.py <dwgm.xlsx> <sttm.xlsx> <out.csv> [gsh_daily.csv] [gas_current.csv]

Gas Supply Hub columns are merged in when fetch_gsh.py has produced them. They join this
series rather than getting tabs of their own because they are the same quantity - an east
coast wholesale gas price in $/GJ - and belong on one axis with the DWGM and the STTM. They
start in 2026 and are blank before that, which is a fact about the source and not a gap in
the build: see fetch_gsh.py for why no deeper history exists.

The master workbooks are republished monthly, so on their own they end at the close of the
last completed month. gas_current.csv carries the same clearing prices from the nightly MIBB
reports (see fetch_gas_current.py) and is used ONLY to extend each column past the last day
the master workbook has for it. The workbook wins wherever both cover a day: it is the
revised, settled version, and letting it win means a later AEMO revision takes effect
instead of being masked by an as-published value. The two are compared on every overlapping
day and the result is printed, so the day they stop agreeing is the day it gets noticed.
"""
import csv, os, sys
from datetime import datetime
from collections import defaultdict
import openpyxl

DWGM, STTM, OUT = sys.argv[1], sys.argv[2], sys.argv[3]
GSH = sys.argv[4] if len(sys.argv) > 4 else None
CURRENT = sys.argv[5] if len(sys.argv) > 5 else None

def as_date(v):
    """Gas_Date arrives as a datetime in some rows and a DD/MM/YYYY string in others."""
    if v is None:
        return None
    if hasattr(v, "date"):
        return v.date().isoformat()
    t = str(v).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d/%m/%y", "%Y/%m/%d"):
        try:
            return datetime.strptime(t[:10], fmt).date().isoformat()
        except ValueError:
            continue
    raise ValueError(f"unparsed gas date: {v!r}")

# --- Victoria DWGM: five scheduling intervals per gas day (06,10,14,18,22) ---
wb = openpyxl.load_workbook(DWGM, read_only=True, data_only=True)
vic = defaultdict(dict)
for gd, hour, price in list(wb["Prices"].iter_rows(values_only=True))[1:]:
    d = as_date(gd)
    if d is None or price is None:
        continue
    try:
        vic[d][int(hour)] = float(price)
    except (TypeError, ValueError):
        continue
wb.close()

# --- STTM hubs: one ex-ante and one ex-post price per gas day per hub ---
wb = openpyxl.load_workbook(STTM, read_only=True, data_only=True)
sttm = defaultdict(dict)
for hub, sheet in (("SYD", "SYD price and withdrawals"),
                   ("ADL", "ADL price and withdrawals"),
                   ("BRI", "BRI price and withdrawals")):
    for row in list(wb[sheet].iter_rows(values_only=True))[1:]:
        d = as_date(row[0])
        if d is None:
            continue
        for label, val in ((f"{hub}_exante", row[1]), (f"{hub}_expost", row[2])):
            try:
                sttm[d][label] = float(val)
            except (TypeError, ValueError):
                pass
wb.close()

# --- Gas Supply Hub: Wallumbilla and SEQ, already one row per gas day ---
gsh, gsh_cols = {}, []
if GSH and os.path.exists(GSH):
    with open(GSH, newline="") as fh:
        r = csv.reader(fh)
        gsh_cols = next(r)[1:]
        for row in r:
            gsh[row[0]] = dict(zip(gsh_cols, row[1:]))

cols = ["VIC_DWGM_6am", "VIC_DWGM_schedule_mean",
        "SYD_exante", "SYD_expost", "ADL_exante", "ADL_expost",
        "BRI_exante", "BRI_expost"]
NCOL = "VIC_DWGM_n_schedules"

# --- the master workbooks, assembled into one row per gas day ---------------
master = {}
for d in sorted(set(vic) | set(sttm)):
    if d < "2007-02-01":
        continue
    v = vic.get(d, {})
    sched = [p for p in v.values() if p is not None]
    s = sttm.get(d, {})
    row = {"VIC_DWGM_6am": v.get(6, ""),
           "VIC_DWGM_schedule_mean": round(sum(sched) / len(sched), 4) if sched else "",
           NCOL: len(sched)}
    row.update({c: s.get(c, "") for c in cols[2:]})
    master[d] = row

# --- the nightly MIBB reports, extending each column past the workbook's end --
current, checked, diffs = {}, 0, []
if CURRENT and os.path.exists(CURRENT):
    with open(CURRENT, newline="") as fh:
        for r in csv.DictReader(fh):
            current[r["date"]] = r
last = {c: max((d for d, row in master.items() if row.get(c) not in ("", None)), default="")
        for c in cols}
for d, row in current.items():
    if d not in master:
        continue
    for c in cols:                       # same day in both: compare, never overwrite
        a, b = master[d].get(c, ""), row.get(c, "")
        if a not in ("", None) and b not in ("", None):
            checked += 1
            if abs(float(a) - float(b)) > 1e-6:
                diffs.append((d, c, float(a), float(b)))
if checked:
    print(f"  MIBB vs master workbook: {checked} values compared, {len(diffs)} differ"
          + (f", max {max(abs(a - b) for _, _, a, b in diffs):.6f}" if diffs else
             " (identical)"), flush=True)
    for d, c, a, b in diffs[:5]:
        print(f"    revised: {d} {c} workbook {a} vs MIBB {b}", flush=True)

dates = sorted(set(master) | set(gsh) | set(current))
with open(OUT, "w", newline="") as fh:
    w = csv.writer(fh)
    w.writerow(["date"] + cols + [NCOL] + gsh_cols)
    for d in dates:
        m = master.get(d, {})
        cur = current.get(d, {})
        out = []
        for c in cols:
            v = m.get(c, "")
            if v in ("", None) and d > last[c]:   # only ever extends, never backfills
                v = cur.get(c, "")
            out.append(v)
        n = m.get(NCOL, "")
        if n in ("", None, 0) and d > last["VIC_DWGM_6am"]:
            n = cur.get(NCOL, "")
        g = gsh.get(d, {})
        w.writerow([d] + out + [n] + [g.get(c, "") for c in gsh_cols])
print("wrote", OUT, "days", len(dates), dates[0], "->", dates[-1],
      f"(GSH on {len(gsh)}, workbook to {max(last.values())})", flush=True)
